"""Ingestion multimédia et fichiers structurés — conversion texte pour IR ARTCB."""

from __future__ import annotations

import csv
import json
import logging
import mimetypes
import subprocess
import tempfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from html.parser import HTMLParser
from io import StringIO
from pathlib import Path
from typing import Any, Final

logger = logging.getLogger("artcb.io.media_ingest")

# --- Extensions supportées (mise en ligne officielle) ---
PLAIN_TEXT_EXTENSIONS: Final[frozenset[str]] = frozenset({
    ".txt", ".md", ".markdown", ".log", ".rst", ".tex", ".sql", ".ini", ".cfg", ".env",
})
JSON_EXTENSIONS: Final[frozenset[str]] = frozenset({".json", ".jsonl", ".ndjson"})
CSV_EXTENSIONS: Final[frozenset[str]] = frozenset({".csv", ".tsv"})
YAML_EXTENSIONS: Final[frozenset[str]] = frozenset({".yaml", ".yml"})
TOML_EXTENSIONS: Final[frozenset[str]] = frozenset({".toml"})
XML_EXTENSIONS: Final[frozenset[str]] = frozenset({".xml", ".rss", ".atom", ".svg"})
HTML_EXTENSIONS: Final[frozenset[str]] = frozenset({".html", ".htm", ".xhtml"})
PDF_EXTENSIONS: Final[frozenset[str]] = frozenset({".pdf"})
IMAGE_EXTENSIONS: Final[frozenset[str]] = frozenset({
    ".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff", ".tif", ".ico",
})
AUDIO_EXTENSIONS: Final[frozenset[str]] = frozenset({
    ".mp3", ".wav", ".ogg", ".m4a", ".flac", ".aac", ".wma", ".opus",
})
VIDEO_EXTENSIONS: Final[frozenset[str]] = frozenset({
    ".mp4", ".mkv", ".webm", ".mov", ".avi", ".wmv", ".flv", ".m4v",
})
DOCX_EXTENSIONS: Final[frozenset[str]] = frozenset({".docx"})
XLSX_EXTENSIONS: Final[frozenset[str]] = frozenset({".xlsx", ".xls", ".ods"})
EPUB_EXTENSIONS: Final[frozenset[str]] = frozenset({".epub"})
RTF_EXTENSIONS: Final[frozenset[str]] = frozenset({".rtf"})
SUBTITLE_EXTENSIONS: Final[frozenset[str]] = frozenset({".srt", ".vtt", ".ass", ".ssa"})

ALL_SUPPORTED_EXTENSIONS: Final[frozenset[str]] = (
    PLAIN_TEXT_EXTENSIONS
    | JSON_EXTENSIONS
    | CSV_EXTENSIONS
    | YAML_EXTENSIONS
    | TOML_EXTENSIONS
    | XML_EXTENSIONS
    | HTML_EXTENSIONS
    | PDF_EXTENSIONS
    | IMAGE_EXTENSIONS
    | AUDIO_EXTENSIONS
    | VIDEO_EXTENSIONS
    | DOCX_EXTENSIONS
    | XLSX_EXTENSIONS
    | EPUB_EXTENSIONS
    | RTF_EXTENSIONS
    | SUBTITLE_EXTENSIONS
)


class MediaIngestError(Exception):
    """Media ingestion failed."""


@dataclass
class IngestedMedia:
    text: str
    media_type: str
    source_path: str
    segments: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def list_supported_formats() -> dict[str, list[str]]:
    """Liste publique des formats pour UI et documentation."""
    return {
        "plain_text": sorted(PLAIN_TEXT_EXTENSIONS),
        "json": sorted(JSON_EXTENSIONS),
        "csv_tsv": sorted(CSV_EXTENSIONS),
        "yaml": sorted(YAML_EXTENSIONS),
        "toml": sorted(TOML_EXTENSIONS),
        "xml": sorted(XML_EXTENSIONS),
        "html": sorted(HTML_EXTENSIONS),
        "pdf": sorted(PDF_EXTENSIONS),
        "image": sorted(IMAGE_EXTENSIONS),
        "audio": sorted(AUDIO_EXTENSIONS),
        "video": sorted(VIDEO_EXTENSIONS),
        "docx": sorted(DOCX_EXTENSIONS),
        "spreadsheet": sorted(XLSX_EXTENSIONS),
        "epub": sorted(EPUB_EXTENSIONS),
        "rtf": sorted(RTF_EXTENSIONS),
        "subtitles": sorted(SUBTITLE_EXTENSIONS),
        "all_extensions": sorted(ALL_SUPPORTED_EXTENSIONS),
    }


def detect_media_type(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in PLAIN_TEXT_EXTENSIONS:
        return "plain_text"
    if ext in JSON_EXTENSIONS:
        return "json"
    if ext in CSV_EXTENSIONS:
        return "csv"
    if ext in YAML_EXTENSIONS:
        return "yaml"
    if ext in TOML_EXTENSIONS:
        return "toml"
    if ext in XML_EXTENSIONS:
        return "xml"
    if ext in HTML_EXTENSIONS:
        return "html"
    if ext in PDF_EXTENSIONS:
        return "pdf"
    if ext in IMAGE_EXTENSIONS:
        return "image"
    if ext in AUDIO_EXTENSIONS:
        return "audio"
    if ext in VIDEO_EXTENSIONS:
        return "video"
    if ext in DOCX_EXTENSIONS:
        return "docx"
    if ext in XLSX_EXTENSIONS:
        return "spreadsheet"
    if ext in EPUB_EXTENSIONS:
        return "epub"
    if ext in RTF_EXTENSIONS:
        return "rtf"
    if ext in SUBTITLE_EXTENSIONS:
        return "subtitles"
    mime, _ = mimetypes.guess_type(str(path))
    if mime:
        if mime.startswith("text/"):
            return "plain_text"
        if "json" in mime:
            return "json"
        if "xml" in mime:
            return "xml"
        if mime.startswith("image/"):
            return "image"
        if mime.startswith("audio/"):
            return "audio"
        if mime.startswith("video/"):
            return "video"
    return "unknown"


def ingest_file(path: Path, *, openai_api_key: str | None = None) -> IngestedMedia:
    """Convertit un fichier local en texte pour le pipeline IR."""
    if not path.is_file():
        raise MediaIngestError(f"Fichier introuvable: {path}")
    media_type = detect_media_type(path)
    warnings: list[str] = []

    if media_type == "plain_text":
        text = _ingest_plain_text(path)
    elif media_type == "json":
        text = _ingest_json(path)
    elif media_type == "csv":
        text = _ingest_csv(path)
    elif media_type == "yaml":
        text = _ingest_yaml(path)
    elif media_type == "toml":
        text = _ingest_toml(path)
    elif media_type == "xml":
        text = _ingest_xml(path)
    elif media_type == "html":
        text = _ingest_html(path)
    elif media_type == "pdf":
        text = _ingest_pdf(path, warnings)
    elif media_type == "image":
        text, warnings_extra = _ingest_image(path, openai_api_key=openai_api_key)
        warnings.extend(warnings_extra)
    elif media_type == "audio":
        text, warnings_extra = _ingest_audio(path, openai_api_key=openai_api_key)
        warnings.extend(warnings_extra)
    elif media_type == "video":
        text, warnings_extra = _ingest_video(path, openai_api_key=openai_api_key)
        warnings.extend(warnings_extra)
    elif media_type == "docx":
        text = _ingest_docx(path)
    elif media_type == "spreadsheet":
        text = _ingest_spreadsheet(path)
    elif media_type == "epub":
        text = _ingest_epub(path)
    elif media_type == "rtf":
        text = _ingest_rtf(path)
    elif media_type == "subtitles":
        text = _ingest_subtitles(path)
    else:
        # Fallback UTF-8
        try:
            raw = path.read_text(encoding="utf-8", errors="strict")
            if raw.strip():
                warnings.append(f"Extension {path.suffix} inconnue — traité comme texte UTF-8")
                text = raw
                media_type = "plain_text"
            else:
                raise MediaIngestError(f"Format non supporté: {path.suffix}")
        except UnicodeDecodeError as exc:
            raise MediaIngestError(
                f"Format non supporté: {path.suffix}. "
                f"Extensions: {', '.join(sorted(ALL_SUPPORTED_EXTENSIONS))}"
            ) from exc

    if not text or not str(text).strip():
        raise MediaIngestError(f"Fichier vide après ingestion: {path.name}")

    return IngestedMedia(
        text=str(text),
        media_type=media_type,
        source_path=str(path),
        warnings=warnings,
    )


def ingest_folder(
    folder: Path,
    *,
    limit: int = 50,
    offset: int = 0,
    extensions: set[str] | None = None,
    openai_api_key: str | None = None,
) -> tuple[str, int, bool]:
    """Lit un dossier — retourne (texte concaténé, count, has_more)."""
    if not folder.is_dir():
        raise MediaIngestError(f"Dossier introuvable: {folder}")
    allowed = extensions or ALL_SUPPORTED_EXTENSIONS
    files = sorted(
        [p for p in folder.rglob("*") if p.is_file() and p.suffix.lower() in allowed],
        key=lambda p: str(p).lower(),
    )
    batch = files[offset : offset + limit]
    parts: list[str] = []
    for fp in batch:
        try:
            ingested = ingest_file(fp, openai_api_key=openai_api_key)
            header = f"--- [{ingested.media_type}] {fp.name} ---"
            parts.append(f"{header}\n{ingested.text}")
            for w in ingested.warnings:
                parts.append(f"[AVERTISSEMENT {fp.name}] {w}")
        except MediaIngestError as exc:
            parts.append(f"--- [ERREUR] {fp.name} ---\n{exc}")
            logger.error("Ingest failed %s: %s", fp, exc)
    text = "\n\n".join(parts)
    has_more = offset + limit < len(files)
    return text, len(batch), has_more


# --- Handlers structurés ---


def _ingest_plain_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _ingest_json(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    ext = path.suffix.lower()
    if ext in (".jsonl", ".ndjson"):
        lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
        objects = []
        for i, ln in enumerate(lines):
            try:
                objects.append(json.loads(ln))
            except json.JSONDecodeError as exc:
                objects.append({"_line": i, "_raw": ln, "_error": str(exc)})
        return json.dumps(objects, ensure_ascii=False, indent=2)
    data = json.loads(raw)
    return json.dumps(data, ensure_ascii=False, indent=2)


def _ingest_csv(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    reader = csv.DictReader(StringIO(raw), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        reader2 = csv.reader(StringIO(raw), delimiter=delimiter)
        rows_plain = list(reader2)
        return "\n".join(" | ".join(cell for cell in row) for row in rows_plain)
    parts = []
    for i, row in enumerate(rows):
        parts.append(f"--- row {i} ---\n{json.dumps(row, ensure_ascii=False)}")
    return "\n\n".join(parts)


def _ingest_yaml(path: Path) -> str:
    try:
        import yaml
    except ImportError as exc:
        raise MediaIngestError("pip install pyyaml — ou: pip install -e '.[media]'") from exc
    data = yaml.safe_load(path.read_text(encoding="utf-8", errors="replace"))
    return json.dumps(data, ensure_ascii=False, indent=2, default=str)


def _ingest_toml(path: Path) -> str:
    try:
        import tomllib
    except ImportError:
        try:
            import tomli as tomllib  # type: ignore[no-redef]
        except ImportError as exc:
            raise MediaIngestError("tomllib (Python 3.11+) ou tomli requis") from exc
    data = tomllib.loads(path.read_text(encoding="utf-8", errors="replace"))
    return json.dumps(data, ensure_ascii=False, indent=2, default=str)


def _ingest_xml(path: Path) -> str:
    tree = ET.parse(path)
    root = tree.getroot()

    def elem_text(el: ET.Element, depth: int = 0) -> str:
        tag = el.tag.split("}")[-1] if "}" in el.tag else el.tag
        text = (el.text or "").strip()
        tail = (el.tail or "").strip()
        children = "\n".join(elem_text(c, depth + 1) for c in el)
        line = f"{'  ' * depth}[{tag}] {text}"
        if children:
            line += f"\n{children}"
        if tail:
            line += f"\n{'  ' * depth}(tail) {tail}"
        return line

    return elem_text(root)


class _HTMLTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._parts: list[str] = []

    def handle_data(self, data: str) -> None:
        s = data.strip()
        if s:
            self._parts.append(s)

    def get_text(self) -> str:
        return "\n".join(self._parts)


def _ingest_html(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    parser = _HTMLTextExtractor()
    parser.feed(raw)
    text = parser.get_text()
    return text if text.strip() else raw


def _ingest_pdf(path: Path, warnings: list[str]) -> str:
    from src.artcb.io.pdf_loader import extract_pdf_text

    text = extract_pdf_text(path)
    if not text.strip():
        warnings.append("PDF sans texte extractible — peut être un scan image")
    return text


def _ingest_spreadsheet(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise MediaIngestError("pip install openpyxl — ou: pip install -e '.[media]'") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"--- sheet: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    parts.append(" | ".join(cells))
        wb.close()
        return "\n".join(parts)
    if ext == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise MediaIngestError("pip install xlrd pour fichiers .xls") from exc
        book = xlrd.open_workbook(str(path))
        parts = []
        for si in range(book.nsheets):
            sh = book.sheet_by_index(si)
            parts.append(f"--- sheet: {sh.name} ---")
            for ri in range(sh.nrows):
                parts.append(" | ".join(str(sh.cell_value(ri, ci)) for ci in range(sh.ncols)))
        return "\n".join(parts)
    raise MediaIngestError(f"Tableur {ext} — convertir en .xlsx ou installer dépendance")


def _ingest_epub(path: Path) -> str:
    try:
        import ebooklib
        from bs4 import BeautifulSoup
        from ebooklib import epub
    except ImportError as exc:
        raise MediaIngestError("pip install ebooklib beautifulsoup4 — ou: pip install -e '.[media]'") from exc
    book = epub.read_epub(str(path))
    parts: list[str] = []
    for item in book.get_items():
        if item.get_type() == ebooklib.ITEM_DOCUMENT:
            soup = BeautifulSoup(item.get_content(), "html.parser")
            text = soup.get_text(separator="\n", strip=True)
            if text:
                parts.append(text)
    return "\n\n".join(parts)


def _ingest_rtf(path: Path) -> str:
    try:
        from striprtf.striprtf import rtf_to_text
    except ImportError:
        raw = path.read_text(encoding="utf-8", errors="replace")
        return raw
    return rtf_to_text(path.read_text(encoding="utf-8", errors="replace"))


def _ingest_subtitles(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    lines: list[str] = []
    for line in raw.splitlines():
        s = line.strip()
        if not s or s.isdigit() or "-->" in s:
            continue
        if s.startswith("WEBVTT"):
            continue
        lines.append(s)
    return "\n".join(lines)


def _ingest_docx(path: Path) -> str:
    try:
        from docx import Document
    except ImportError as exc:
        raise MediaIngestError("pip install python-docx — ou: pip install -e '.[media]'") from exc
    doc = Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _ingest_image(path: Path, *, openai_api_key: str | None) -> tuple[str, list[str]]:
    warnings: list[str] = []
    try:
        import pytesseract
        from PIL import Image

        text = pytesseract.image_to_string(Image.open(path))
        if text.strip():
            return f"[OCR]\n{text.strip()}", warnings
        warnings.append("OCR Tesseract vide")
    except ImportError:
        warnings.append("pytesseract/Pillow non installés — pip install -e '.[media]'")
    except Exception as exc:
        warnings.append(f"OCR échoué: {exc}")

    if openai_api_key:
        try:
            import base64

            import httpx

            b64 = base64.standard_b64encode(path.read_bytes()).decode("ascii")
            mime = mimetypes.guess_type(str(path))[0] or "image/jpeg"
            with httpx.Client(timeout=90.0) as client:
                r = client.post(
                    "https://api.openai.com/v1/chat/completions",
                    headers={"Authorization": f"Bearer {openai_api_key}"},
                    json={
                        "model": "gpt-4o-mini",
                        "messages": [{
                            "role": "user",
                            "content": [
                                {
                                    "type": "text",
                                    "text": "Décris cette image en français pour mémorisation cognitive. Inclus tous les textes visibles.",
                                },
                                {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
                            ],
                        }],
                        "max_tokens": 2000,
                    },
                )
                r.raise_for_status()
                desc = r.json()["choices"][0]["message"]["content"]
                return f"[VISION]\n{desc}", warnings
        except Exception as exc:
            warnings.append(f"Vision API échouée: {exc}")

    raise MediaIngestError(
        "Image non lisible — pip install -e '.[media]' (OCR) ou connectez OpenAI pour vision"
    )


def _ingest_audio(path: Path, *, openai_api_key: str | None) -> tuple[str, list[str]]:
    warnings: list[str] = []
    if openai_api_key:
        try:
            import httpx

            with path.open("rb") as f, httpx.Client(timeout=120.0) as client:
                r = client.post(
                    "https://api.openai.com/v1/audio/transcriptions",
                    headers={"Authorization": f"Bearer {openai_api_key}"},
                    files={"file": (path.name, f, "application/octet-stream")},
                    data={"model": "whisper-1", "language": "fr"},
                )
                r.raise_for_status()
                transcript = r.json().get("text", "")
                if transcript.strip():
                    return f"[TRANSCRIPT]\n{transcript.strip()}", warnings
        except Exception as exc:
            warnings.append(f"Whisper API échouée: {exc}")

    try:
        subprocess.run(
            ["whisper", str(path), "--language", "French", "--output_format", "txt", "--fp16", "False"],
            capture_output=True,
            text=True,
            timeout=300,
            check=False,
        )
        txt_path = path.with_suffix(".txt")
        if txt_path.is_file():
            return f"[TRANSCRIPT-LOCAL]\n{txt_path.read_text(encoding='utf-8')}", warnings
    except FileNotFoundError:
        warnings.append("whisper CLI non installé")
    except Exception as exc:
        warnings.append(f"whisper local: {exc}")

    raise MediaIngestError(
        "Audio non transcrit — connectez OpenAI (Whisper API) ou installez openai-whisper CLI"
    )


def _ingest_video(path: Path, *, openai_api_key: str | None) -> tuple[str, list[str]]:
    warnings: list[str] = []
    with tempfile.TemporaryDirectory() as tmp:
        audio_path = Path(tmp) / "extracted.wav"
        try:
            subprocess.run(
                [
                    "ffmpeg", "-y", "-i", str(path),
                    "-vn", "-acodec", "pcm_s16le", "-ar", "16000", str(audio_path),
                ],
                capture_output=True,
                timeout=120,
                check=True,
            )
        except FileNotFoundError:
            raise MediaIngestError("ffmpeg requis pour vidéo — apt install ffmpeg") from None
        except subprocess.CalledProcessError as exc:
            err = exc.stderr[:200].decode() if isinstance(exc.stderr, bytes) else str(exc.stderr)[:200]
            raise MediaIngestError(f"ffmpeg échec extraction audio: {err}") from exc

        if audio_path.is_file() and audio_path.stat().st_size > 0:
            transcript, w = _ingest_audio(audio_path, openai_api_key=openai_api_key)
            warnings.extend(w)
            meta = {"source_video": str(path), "audio_extracted": True}
            return f"{transcript}\n\n[META]{json.dumps(meta)}", warnings

    raise MediaIngestError("Vidéo sans piste audio extractible")
